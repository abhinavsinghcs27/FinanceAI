from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import secrets
import sqlite3
import urllib.error
import urllib.request
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Annotated, Any

from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency fallback
    Workbook = None
    load_workbook = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except ImportError:  # pragma: no cover - optional dependency fallback
    letter = None
    canvas = None


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("FINANCEAI_DB_PATH", BASE_DIR / "financeai.db"))
AI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5.2")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
SESSION_TTL_HOURS = int(os.getenv("SESSION_TTL_HOURS", "24"))
REPORTS_DIR = Path(os.getenv("FINANCEAI_REPORTS_DIR", BASE_DIR / "generated_reports"))
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")
REPORT_FROM_EMAIL = os.getenv("REPORT_FROM_EMAIL", "reports@financeai.local")
ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "ALLOWED_ORIGINS",
        "http://localhost:3000,http://127.0.0.1:3000,http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if origin.strip()
]
SEED_DEMO_DATA = os.getenv("SEED_DEMO_DATA", "true").lower() not in {"0", "false", "no"}
AI_DISCLAIMER = (
    "AI insights are educational and are not financial, investment, tax, or legal advice."
)
TEMPLATES = [
    {"title": "Monthly Summary", "subtitle": "Performance and holdings"},
    {"title": "Tax Report", "subtitle": "Capital gains and losses"},
    {"title": "Quarterly Review", "subtitle": "Comprehensive analysis"},
    {"title": "Transaction Export", "subtitle": "All trades and activity"},
]
SCHEDULED_REPORT = {
    "title": "Monthly Performance",
    "subtitle": "Every 1st of the month",
    "status": "Active",
}
UPLOAD_COLUMNS = [
    "Date (YYYY-MM-DD)",
    "Type (Buy/Sell)",
    "Symbol (Stock Ticker)",
    "Quantity (Number of shares)",
    "Price (Price per share)",
]
SUPPORTED_BROKERAGES = [
    "Robinhood",
    "E*TRADE",
    "TD Ameritrade",
    "Fidelity",
    "Charles Schwab",
    "Interactive Brokers",
]
PRICE_BOOK: dict[str, dict[str, Any]] = {
    "AAPL": {"company": "Apple Inc.", "price": 198.30, "sector": "Technology", "risk": "Medium Risk"},
    "MSFT": {"company": "Microsoft Corporation", "price": 421.10, "sector": "Technology", "risk": "Low Risk"},
    "NVDA": {"company": "NVIDIA Corporation", "price": 972.40, "sector": "Technology", "risk": "Medium Risk"},
    "AMD": {"company": "Advanced Micro Devices", "price": 176.40, "sector": "Technology", "risk": "Medium Risk"},
    "GOOGL": {"company": "Alphabet Inc.", "price": 164.20, "sector": "Technology", "risk": "Medium Risk"},
    "JNJ": {"company": "Johnson & Johnson", "price": 163.10, "sector": "Healthcare", "risk": "Low Risk"},
    "JPM": {"company": "JPMorgan Chase & Co.", "price": 201.20, "sector": "Finance", "risk": "Low Risk"},
    "XOM": {"company": "Exxon Mobil Corporation", "price": 121.60, "sector": "Energy", "risk": "Medium Risk"},
    "KO": {"company": "The Coca-Cola Company", "price": 63.20, "sector": "Consumer", "risk": "Low Risk"},
    "BND": {"company": "Vanguard Total Bond Market ETF", "price": 72.80, "sector": "Bonds", "risk": "Low Risk"},
    "VOO": {"company": "Vanguard S&P 500 ETF", "price": 512.70, "sector": "ETFs", "risk": "Low Risk"},
}
DEFAULT_TRANSACTIONS: list[dict[str, Any]] = [
    {"date": "2026-01-10", "type": "Buy", "symbol": "AAPL", "quantity": 20.0, "price": 188.50},
    {"date": "2026-01-12", "type": "Buy", "symbol": "MSFT", "quantity": 10.0, "price": 405.25},
    {"date": "2026-01-18", "type": "Buy", "symbol": "NVDA", "quantity": 8.0, "price": 910.00},
    {"date": "2026-02-02", "type": "Buy", "symbol": "JNJ", "quantity": 12.0, "price": 158.10},
    {"date": "2026-02-11", "type": "Buy", "symbol": "BND", "quantity": 50.0, "price": 71.25},
    {"date": "2026-03-04", "type": "Buy", "symbol": "VOO", "quantity": 15.0, "price": 501.40},
    {"date": "2026-03-11", "type": "Buy", "symbol": "JPM", "quantity": 14.0, "price": 194.10},
    {"date": "2026-03-18", "type": "Sell", "symbol": "AAPL", "quantity": 4.0, "price": 196.20},
]


app = FastAPI(
    title="FinanceAI API",
    version="2.0.0",
    description="SQLite-backed MVP backend for the FinanceAI frontend.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class SignupRequest(BaseModel):
    full_name: str = Field(min_length=2)
    email: EmailStr
    password: str = Field(min_length=6)
    goal: str = "Long-term growth"


class LoginRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)


class ProfileUpdateRequest(BaseModel):
    full_name: str = Field(min_length=2)
    email: EmailStr
    risk_preference: str
    primary_goal: str
    base_currency: str
    reports: str


class ReportGenerateRequest(BaseModel):
    format: str
    date_range: str
    sections: list[str]


class EmailReportRequest(BaseModel):
    email: EmailStr
    report_name: str


class AIChatRequest(BaseModel):
    question: str = Field(min_length=3, max_length=800)


class AIReportInsightRequest(BaseModel):
    format: str = "PDF Document"
    date_range: str = "Last Month"
    sections: list[str] = Field(default_factory=list)


class TransactionRequest(BaseModel):
    date: str
    type: str
    symbol: str
    quantity: float = Field(gt=0)
    price: float = Field(gt=0)


def utc_now_iso() -> str:
    return datetime.now(UTC).isoformat()


def get_db() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def parse_date(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d")


def month_label(value: datetime) -> str:
    return value.strftime("%b")


def currency(amount: float) -> str:
    return f"${amount:,.2f}"


def percent(value: float) -> str:
    prefix = "+" if value > 0 else ""
    return f"{prefix}{value:.1f}%"


def hash_password(password: str, salt: str) -> str:
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000)
    return digest.hex()


def make_password_fields(password: str) -> tuple[str, str]:
    salt = secrets.token_hex(16)
    return salt, hash_password(password, salt)


def verify_password(password: str, salt: str, expected_hash: str) -> bool:
    return hmac.compare_digest(hash_password(password, salt), expected_hash)


def create_session(connection: sqlite3.Connection, user_id: int) -> str:
    token = secrets.token_urlsafe(32)
    connection.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
    connection.execute(
        "INSERT INTO sessions (token, user_id, created_at) VALUES (?, ?, ?)",
        (token, user_id, utc_now_iso()),
    )
    connection.commit()
    return token


def parse_iso_datetime(value: str) -> datetime:
    parsed = datetime.fromisoformat(value)
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)


def resolve_asset(symbol: str) -> dict[str, Any]:
    asset = PRICE_BOOK.get(symbol.upper())
    if asset:
        return asset
    return {
        "company": f"{symbol.upper()} Holdings",
        "price": 100.0,
        "sector": "Other",
        "risk": "Medium Risk",
    }


def public_user_profile(user: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": user["id"],
        "full_name": user["full_name"],
        "email": user["email"],
        "plan": user["plan"],
        "member_since": user["member_since"],
        "risk_preference": user["risk_preference"],
        "primary_goal": user["primary_goal"],
        "base_currency": user["base_currency"],
        "reports": user["reports"],
    }


def current_user(
    authorization: Annotated[str | None, Header()] = None,
) -> dict[str, Any]:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authentication required.")

    token = authorization.removeprefix("Bearer ").strip()
    with get_db() as connection:
        row = connection.execute(
            """
            SELECT users.*, sessions.created_at AS session_created_at
            FROM sessions
            JOIN users ON users.id = sessions.user_id
            WHERE sessions.token = ?
            """,
            (token,),
        ).fetchone()

        if row:
            session_age = datetime.now(UTC) - parse_iso_datetime(row["session_created_at"])
            if session_age > timedelta(hours=SESSION_TTL_HOURS):
                connection.execute("DELETE FROM sessions WHERE token = ?", (token,))
                connection.commit()
                row = None

    if not row:
        raise HTTPException(status_code=401, detail="Invalid or expired session.")

    return {"token": token, "user": row}


def fetch_transactions(connection: sqlite3.Connection, user_id: int) -> list[sqlite3.Row]:
    return connection.execute(
        """
        SELECT date, type, symbol, quantity, price
        FROM transactions
        WHERE user_id = ?
        ORDER BY date ASC, id ASC
        """,
        (user_id,),
    ).fetchall()


def build_positions(connection: sqlite3.Connection, user_id: int) -> dict[str, dict[str, Any]]:
    positions: dict[str, dict[str, Any]] = {}
    for transaction in fetch_transactions(connection, user_id):
        symbol = transaction["symbol"].upper()
        asset = resolve_asset(symbol)
        position = positions.setdefault(
            symbol,
            {
                "symbol": symbol,
                "quantity": 0.0,
                "cost_basis": 0.0,
                "buys": 0.0,
                "sells": 0.0,
                "company": asset["company"],
                "sector": asset["sector"],
                "risk": asset["risk"],
            },
        )

        quantity = float(transaction["quantity"])
        trade_value = quantity * float(transaction["price"])
        if transaction["type"].lower() == "buy":
            position["quantity"] += quantity
            position["cost_basis"] += trade_value
            position["buys"] += trade_value
        else:
            average_cost = (
                position["cost_basis"] / position["quantity"] if position["quantity"] > 0 else 0.0
            )
            position["quantity"] = max(position["quantity"] - quantity, 0.0)
            position["cost_basis"] = max(position["cost_basis"] - average_cost * quantity, 0.0)
            position["sells"] += trade_value

    return {symbol: item for symbol, item in positions.items() if item["quantity"] > 0}


def position_snapshot(connection: sqlite3.Connection, user: sqlite3.Row) -> list[dict[str, Any]]:
    holdings = []
    for symbol, position in build_positions(connection, user["id"]).items():
        live = resolve_asset(symbol)
        market_value = position["quantity"] * float(live["price"])
        average_cost = position["cost_basis"] / position["quantity"]
        gain_loss = market_value - position["cost_basis"]
        holdings.append(
            {
                **position,
                "company": live["company"],
                "sector": live["sector"],
                "current_price": float(live["price"]),
                "market_value": market_value,
                "average_cost": average_cost,
                "gain_loss": gain_loss,
            }
        )
    return sorted(holdings, key=lambda item: item["market_value"], reverse=True)


def total_portfolio_value(holdings: list[dict[str, Any]], user: sqlite3.Row) -> float:
    return sum(item["market_value"] for item in holdings) + float(user["cash_balance"])


def allocation_breakdown(holdings: list[dict[str, Any]], user: sqlite3.Row) -> list[dict[str, Any]]:
    total_value = total_portfolio_value(holdings, user)
    buckets = {"Stocks": 0.0, "Bonds": 0.0, "ETFs": 0.0, "Cash": float(user["cash_balance"])}
    for holding in holdings:
        if holding["sector"] == "Bonds":
            buckets["Bonds"] += holding["market_value"]
        elif holding["sector"] == "ETFs":
            buckets["ETFs"] += holding["market_value"]
        else:
            buckets["Stocks"] += holding["market_value"]

    colors = {"Stocks": "#2563eb", "Bonds": "#10b981", "ETFs": "#f59e0b", "Cash": "#64748b"}
    return [
        {
            "label": label,
            "value": round((amount / total_value) * 100) if total_value else 0,
            "color": colors[label],
        }
        for label, amount in buckets.items()
    ]


def performance_series(connection: sqlite3.Connection, user_id: int) -> list[dict[str, Any]]:
    transactions = fetch_transactions(connection, user_id)
    monthly_buys: dict[str, float] = {}
    monthly_sells: dict[str, float] = {}
    for transaction in transactions:
        month = month_label(parse_date(transaction["date"]))
        bucket = monthly_buys if transaction["type"].lower() == "buy" else monthly_sells
        bucket[month] = bucket.get(month, 0.0) + float(transaction["quantity"]) * float(transaction["price"])

    months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    running = 42.0
    series = []
    for index, month in enumerate(months):
        running += monthly_buys.get(month, 0.0) / 1000
        running -= monthly_sells.get(month, 0.0) / 1400
        running += 1.5 + (index % 3) * 0.9
        series.append({"month": month, "value": round(running, 1)})
    return series


def dashboard_payload(connection: sqlite3.Connection, user: sqlite3.Row) -> dict[str, Any]:
    holdings = position_snapshot(connection, user)
    invested_value = sum(item["market_value"] for item in holdings)
    cost_basis = sum(item["cost_basis"] for item in holdings)
    total_value = total_portfolio_value(holdings, user)
    total_gain = invested_value - cost_basis
    day_gain = invested_value * 0.0068
    best_performer = holdings[0]["symbol"] if holdings else "Cash"
    cash_share = (float(user["cash_balance"]) / total_value * 100) if total_value else 0.0
    growth_pct = (total_gain / cost_basis * 100) if cost_basis else 0.0

    return {
        "summary_cards": [
            {
                "title": "Total Portfolio Value",
                "value": currency(total_value),
                "change": percent(growth_pct),
                "detail": f"Across {len(holdings)} active holdings",
                "tone": "#0f9f6e" if total_gain >= 0 else "#dc2626",
            },
            {
                "title": "Total Gain/Loss",
                "value": currency(total_gain),
                "change": percent(growth_pct),
                "detail": f"Best performer: {best_performer}",
                "tone": "#0f9f6e" if total_gain >= 0 else "#dc2626",
            },
            {
                "title": "Day's Gain/Loss",
                "value": currency(day_gain),
                "change": "+0.7%",
                "detail": "Based on latest marked prices",
                "tone": "#2563eb",
            },
            {
                "title": "Cash Available",
                "value": currency(float(user["cash_balance"])),
                "change": f"{cash_share:.0f}%",
                "detail": "Ready for reallocation",
                "tone": "#b45309",
            },
        ],
        "performance": performance_series(connection, user["id"]),
        "allocation": allocation_breakdown(holdings, user),
    }


def sector_exposure(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    invested = sum(item["market_value"] for item in holdings)
    totals: dict[str, float] = {}
    for holding in holdings:
        totals[holding["sector"]] = totals.get(holding["sector"], 0.0) + holding["market_value"]

    sectors = [
        {"label": label, "value": round(amount / invested * 100) if invested else 0}
        for label, amount in sorted(totals.items(), key=lambda item: item[1], reverse=True)
    ]
    if len(sectors) < 6:
        sectors.append({"label": "Other", "value": max(0, 100 - sum(item["value"] for item in sectors))})
    return sectors[:6]


def risk_payload(connection: sqlite3.Connection, user: sqlite3.Row) -> dict[str, Any]:
    holdings = position_snapshot(connection, user)
    total_value = total_portfolio_value(holdings, user)
    invested = sum(item["market_value"] for item in holdings)
    allocation = allocation_breakdown(holdings, user)
    sectors = sector_exposure(holdings)
    largest_position = max((item["market_value"] for item in holdings), default=0.0)
    largest_sector = max((item["value"] for item in sectors), default=0)
    cash_share = next((item["value"] for item in allocation if item["label"] == "Cash"), 0)

    diversification = max(35, 100 - int(largest_position / total_value * 100)) if total_value else 50
    sector_concentration = max(20, min(95, largest_sector + 25))
    volatility = min(95, 30 + len([item for item in holdings if item["risk"] != "Low Risk"]) * 8)
    market_exposure = min(95, max(25, int(invested / total_value * 100) + 5)) if total_value else 40
    liquidity = min(95, max(25, cash_share + 35))
    currency_risk = 45 if user["base_currency"] == "USD" else 60
    overall = round(
        (
            volatility
            + (100 - diversification)
            + sector_concentration
            + market_exposure
            + (100 - liquidity)
            + currency_risk
        )
        / 6
    )

    label = "Low Risk" if overall < 45 else "Medium Risk" if overall < 70 else "High Risk"
    benchmark = max(40, overall - 8)
    trend = [
        {
            "month": month,
            "portfolio": max(25, min(90, overall - 8 + index * 2)),
            "benchmark": max(25, min(85, benchmark - 5 + index * 2)),
        }
        for index, month in enumerate(["Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb"])
    ]

    alerts = []
    if largest_sector >= 35:
        alerts.append(
            {
                "title": "High Concentration Alert",
                "message": f"Your largest sector represents {largest_sector}% of invested assets. Consider diversifying.",
                "tone": "warning",
            }
        )
    if cash_share < 8:
        alerts.append(
            {
                "title": "Low Cash Buffer",
                "message": "Available cash is thin for new opportunities or market pullbacks.",
                "tone": "info",
            }
        )
    if not alerts:
        alerts.append(
            {
                "title": "Risk Profile Stable",
                "message": "Your allocation is within a manageable range for the selected investor profile.",
                "tone": "info",
            }
        )

    recommendations = []
    if largest_sector >= 35:
        recommendations.append(
            {
                "title": "Reduce Top Sector Exposure",
                "description": "Trim the most concentrated sector and rebalance toward underweighted areas.",
                "color": "#2563eb",
                "bg": "#dbeafe",
                "icon": "shield",
            }
        )
    if cash_share < 12:
        recommendations.append(
            {
                "title": "Rebuild Cash Reserve",
                "description": "Target a 10-15% cash allocation so you can absorb volatility and deploy capital selectively.",
                "color": "#16a34a",
                "bg": "#dcfce7",
                "icon": "trendDown",
            }
        )
    recommendations.append(
        {
            "title": "Use Watchlist for Staged Entries",
            "description": "Phase into new ideas over several trades instead of adding all exposure at once.",
            "color": "#9333ea",
            "bg": "#f3e8ff",
            "icon": "line",
        }
    )

    return {
        "score": {
            "overall": overall,
            "industry_average": 55,
            "recommended_range": "45-60",
            "label": label,
        },
        "radar": [
            {"label": "Volatility", "value": volatility},
            {"label": "Diversification", "value": diversification},
            {"label": "Sector Concentration", "value": sector_concentration},
            {"label": "Market Exposure", "value": market_exposure},
            {"label": "Liquidity", "value": liquidity},
            {"label": "Currency Risk", "value": currency_risk},
        ],
        "trend": trend,
        "sectors": sectors,
        "detail_metrics": [
            {"label": "Volatility", "value": volatility},
            {"label": "Diversification", "value": diversification},
            {"label": "Sector Concentration", "value": sector_concentration},
            {"label": "Market Exposure", "value": market_exposure},
            {"label": "Liquidity", "value": liquidity},
        ],
        "alerts": alerts,
        "recommendations": recommendations,
    }


def fetch_watchlist(connection: sqlite3.Connection, user_id: int) -> list[str]:
    rows = connection.execute(
        "SELECT symbol FROM watchlist WHERE user_id = ? ORDER BY created_at DESC, id DESC",
        (user_id,),
    ).fetchall()
    return [row["symbol"] for row in rows]


def ai_context(connection: sqlite3.Connection, user: sqlite3.Row) -> dict[str, Any]:
    holdings = position_snapshot(connection, user)
    risk = risk_payload(connection, user)
    dashboard = dashboard_payload(connection, user)
    recommendations = recommendations_payload(connection, user)
    return {
        "user": {
            "risk_preference": user["risk_preference"],
            "primary_goal": user["primary_goal"],
            "base_currency": user["base_currency"],
            "cash_balance": float(user["cash_balance"]),
        },
        "portfolio": {
            "total_value": total_portfolio_value(holdings, user),
            "holdings": [
                {
                    "symbol": item["symbol"],
                    "company": item["company"],
                    "sector": item["sector"],
                    "quantity": round(item["quantity"], 4),
                    "market_value": round(item["market_value"], 2),
                    "gain_loss": round(item["gain_loss"], 2),
                    "risk": item["risk"],
                }
                for item in holdings
            ],
            "allocation": dashboard["allocation"],
        },
        "risk": risk,
        "recommendations": recommendations["stocks"],
    }


def rule_based_ai_insight(kind: str, context: dict[str, Any], question: str | None = None) -> dict[str, Any]:
    holdings = context["portfolio"]["holdings"]
    allocation = context["portfolio"]["allocation"]
    risk = context["risk"]["score"]
    largest_holding = max(holdings, key=lambda item: item["market_value"], default=None)
    largest_allocation = max(allocation, key=lambda item: item["value"], default={"label": "Cash", "value": 0})
    cash = next((item for item in allocation if item["label"] == "Cash"), {"value": 0})

    top_actions = []
    if largest_allocation["value"] >= 45:
        top_actions.append(f"Review concentration in {largest_allocation['label']} before adding more exposure.")
    if cash["value"] < 10:
        top_actions.append("Rebuild a 10-15% cash buffer before taking larger new positions.")
    if largest_holding:
        top_actions.append(f"Stress test {largest_holding['symbol']} because it is the largest position by value.")
    top_actions.append("Use staged entries for new investments instead of deploying all cash at once.")

    summary = (
        f"Your portfolio is currently rated {risk['label']} with a score of {risk['overall']}/100. "
        f"The largest allocation bucket is {largest_allocation['label']} at {largest_allocation['value']}%."
    )
    if question:
        summary = f"For your question, '{question}', the key issue is whether the portfolio still matches your stated goal and risk preference. {summary}"

    return {
        "source": "local_rules",
        "summary": summary,
        "key_points": [
            f"Risk score: {risk['overall']}/100 versus recommended range {risk['recommended_range']}.",
            f"Cash allocation: {cash['value']}%.",
            f"Active holdings reviewed: {len(holdings)}.",
        ],
        "actions": top_actions[:4],
        "disclaimer": AI_DISCLAIMER,
    }


def call_openai_insight(kind: str, context: dict[str, Any], question: str | None = None) -> dict[str, Any]:
    if not OPENAI_API_KEY:
        return rule_based_ai_insight(kind, context, question)

    prompt = {
        "task": kind,
        "question": question,
        "context": context,
        "required_json_schema": {
            "source": "openai",
            "summary": "short paragraph",
            "key_points": ["3 concise bullets"],
            "actions": ["3 to 5 concrete next steps"],
            "disclaimer": AI_DISCLAIMER,
        },
    }
    body = json.dumps(
        {
            "model": AI_MODEL,
            "instructions": (
                "You are FinanceAI, an educational portfolio analysis assistant. "
                "Do not give guaranteed returns or personalized licensed financial advice. "
                "Be specific to the provided portfolio data. Return valid JSON only."
            ),
            "input": json.dumps(prompt),
            "max_output_tokens": 700,
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        OPENAI_RESPONSES_URL,
        data=body,
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = json.loads(response.read().decode("utf-8"))
        text = raw.get("output_text", "")
        if not text:
            output = raw.get("output", [])
            text = "".join(
                content.get("text", "")
                for item in output
                for content in item.get("content", [])
                if content.get("type") == "output_text"
            )
        parsed = json.loads(text)
        parsed.setdefault("source", "openai")
        parsed.setdefault("disclaimer", AI_DISCLAIMER)
        return parsed
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, KeyError):
        fallback = rule_based_ai_insight(kind, context, question)
        fallback["source"] = "local_rules_after_ai_error"
        return fallback


def recommendation_description(sector: str, goal: str, is_watchlist: bool) -> str:
    if "growth" in goal.lower() and sector == "Technology":
        base = "Fits a growth-oriented portfolio with strong upside and durable product demand."
    elif "income" in goal.lower():
        base = "Supports a steadier income profile with more defensive return characteristics."
    else:
        base = "Improves diversification while keeping the portfolio aligned with your target profile."
    if is_watchlist:
        return f"{base} Already on your watchlist, making it a natural next candidate to monitor."
    return base


def recommendations_payload(connection: sqlite3.Connection, user: sqlite3.Row) -> dict[str, Any]:
    holdings = position_snapshot(connection, user)
    owned_symbols = {item["symbol"] for item in holdings}
    sectors_owned = {item["sector"] for item in holdings}
    watchlist = set(fetch_watchlist(connection, user["id"]))
    ideas = []
    risk_profile = user["risk_preference"].lower()
    goal = user["primary_goal"].lower()

    for ticker, asset in PRICE_BOOK.items():
        price = float(asset["price"])
        score = 68
        if ticker in watchlist:
            score += 8
        if ticker not in owned_symbols:
            score += 10
        if asset["sector"] not in sectors_owned:
            score += 7
        if "growth" in goal and asset["sector"] == "Technology":
            score += 8
        if "income" in goal and asset["sector"] in {"Bonds", "Consumer", "Healthcare"}:
            score += 8
        if risk_profile == "conservative" and asset["risk"] == "Low Risk":
            score += 7
        if risk_profile == "aggressive" and asset["risk"] != "Low Risk":
            score += 6

        upside_pct = round(min(18.0, max(4.5, score / 7)), 1)
        action = "Strong Buy" if score >= 88 else "Buy" if score >= 78 else "Watch"
        confidence = "Very High" if score >= 90 else "High" if score >= 82 else "Moderate"
        change = round(((score % 7) - 2) * 0.7, 1)
        ideas.append(
            {
                "ticker": ticker,
                "company": asset["company"],
                "price": currency(price),
                "change": f"{abs(change):.1f}%",
                "change_color": "#16a34a" if change >= 0 else "#dc2626",
                "action": action,
                "risk": asset["risk"],
                "ai_score": min(score, 96),
                "description": recommendation_description(asset["sector"], user["primary_goal"], ticker in watchlist),
                "target_price": currency(price * (1 + upside_pct / 100)),
                "upside": f"+{upside_pct:.1f}%",
                "confidence_label": confidence,
            }
        )

    ideas.sort(key=lambda item: item["ai_score"], reverse=True)
    diversification_fit = min(95, 55 + len(PRICE_BOOK) - len(sectors_owned) * 3)
    return {
        "summary": {
            "opportunities_found": len(PRICE_BOOK),
            "message": f"{len(watchlist)} tickers are already on your watchlist.",
        },
        "filters": ["AI Picks", "Trending", "High Dividend"],
        "stocks": ideas[:3],
        "fit_analysis": [
            {"label": "Risk Alignment", "value": 82 if user["risk_preference"] != "Conservative" else 90},
            {"label": "Sector Diversification", "value": max(55, min(92, diversification_fit))},
            {"label": "Growth Potential", "value": 88 if "growth" in goal else 72},
            {"label": "Dividend Yield", "value": 70 if "income" in goal else 48},
            {"label": "Valuation", "value": 68},
        ],
    }


def reports_payload(connection: sqlite3.Connection, user_id: int) -> dict[str, Any]:
    recent_rows = connection.execute(
        """
        SELECT id, title, meta, content_type
        FROM reports
        WHERE user_id = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 8
        """,
        (user_id,),
    ).fetchall()
    return {
        "templates": TEMPLATES,
        "recent": [dict(row) for row in recent_rows],
        "scheduled": SCHEDULED_REPORT,
    }


def uploads_payload(connection: sqlite3.Connection, user_id: int) -> dict[str, Any]:
    recent_rows = connection.execute(
        """
        SELECT name, date_label AS date, status
        FROM uploads
        WHERE user_id = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 8
        """,
        (user_id,),
    ).fetchall()
    return {
        "required_columns": UPLOAD_COLUMNS,
        "supported_brokerages": SUPPORTED_BROKERAGES,
        "recent_uploads": [dict(row) for row in recent_rows],
    }


def coerce_transaction(row: dict[str, str]) -> dict[str, Any]:
    normalized = {key.strip().lower(): str(value).strip() for key, value in row.items() if key and value is not None}
    if not normalized:
        raise ValueError("Encountered an empty row.")
    try:
        date_value = normalized["date"]
        parse_date(date_value)
        tx_type = normalized["type"].capitalize()
        symbol = normalized["symbol"].upper()
        quantity = float(normalized["quantity"])
        price = float(normalized["price"])
    except KeyError as exc:
        raise ValueError(f"Missing required column: {exc.args[0]}") from exc
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    if tx_type not in {"Buy", "Sell"}:
        raise ValueError("Type must be Buy or Sell.")
    if quantity <= 0 or price <= 0:
        raise ValueError("Quantity and price must be positive values.")

    return {
        "date": date_value,
        "type": tx_type,
        "symbol": symbol,
        "quantity": quantity,
        "price": price,
    }


def parse_excel_transactions(contents: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("Excel parsing is not installed. Install openpyxl or upload CSV.")

    workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel file must include a header row and at least one transaction.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed = []
    for raw_row in rows[1:]:
        item = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        if any(value is not None and str(value).strip() for value in item.values()):
            parsed.append(coerce_transaction(item))

    if not parsed:
        raise ValueError("Excel file did not contain any transaction rows.")
    return parsed


def transaction_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "date": row["date"],
        "type": row["type"],
        "symbol": row["symbol"],
        "quantity": row["quantity"],
        "price": row["price"],
        "source": row["source"],
        "trade_value": round(float(row["quantity"]) * float(row["price"]), 2),
    }


def validate_incoming_transactions(
    connection: sqlite3.Connection,
    user: sqlite3.Row,
    parsed_rows: list[dict[str, Any]],
) -> None:
    running_cash = float(user["cash_balance"])
    running_positions = {
        symbol: values["quantity"] for symbol, values in build_positions(connection, user["id"]).items()
    }

    for transaction in sorted(parsed_rows, key=lambda item: item["date"]):
        symbol = transaction["symbol"]
        quantity = float(transaction["quantity"])
        trade_value = quantity * float(transaction["price"])
        if transaction["type"] == "Buy":
            if trade_value > running_cash:
                raise ValueError(f"Insufficient cash balance for buying {symbol}.")
            running_cash -= trade_value
            running_positions[symbol] = running_positions.get(symbol, 0.0) + quantity
        else:
            if quantity > running_positions.get(symbol, 0.0):
                raise ValueError(f"Cannot sell {quantity} shares of {symbol}; holdings are too low.")
            running_positions[symbol] -= quantity
            running_cash += trade_value


def insert_transactions(
    connection: sqlite3.Connection,
    user: sqlite3.Row,
    parsed_rows: list[dict[str, Any]],
    filename: str,
) -> dict[str, Any]:
    validate_incoming_transactions(connection, user, parsed_rows)

    cash_balance = float(user["cash_balance"])
    for transaction in parsed_rows:
        trade_value = float(transaction["quantity"]) * float(transaction["price"])
        if transaction["type"] == "Buy":
            cash_balance -= trade_value
        else:
            cash_balance += trade_value

        connection.execute(
            """
            INSERT INTO transactions (user_id, date, type, symbol, quantity, price, source)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user["id"],
                transaction["date"],
                transaction["type"],
                transaction["symbol"],
                transaction["quantity"],
                transaction["price"],
                filename,
            ),
        )

    connection.execute(
        "UPDATE users SET cash_balance = ? WHERE id = ?",
        (cash_balance, user["id"]),
    )
    connection.execute(
        """
        INSERT INTO uploads (user_id, name, date_label, status, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            user["id"],
            filename,
            f"Uploaded {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}",
            "Processed",
            utc_now_iso(),
        ),
    )
    connection.commit()

    return {
        "message": f"Uploaded {len(parsed_rows)} transactions and refreshed portfolio analytics.",
        "filename": filename,
        "columns": ["Date", "Type", "Symbol", "Quantity", "Price"],
        "row_count": len(parsed_rows),
        "preview": [
            {
                "Date": item["date"],
                "Type": item["type"],
                "Symbol": item["symbol"],
                "Quantity": item["quantity"],
                "Price": item["price"],
            }
            for item in parsed_rows[:3]
        ],
    }


def insert_single_transaction(
    connection: sqlite3.Connection,
    user: sqlite3.Row,
    payload: TransactionRequest,
) -> dict[str, Any]:
    transaction = coerce_transaction(
        {
            "date": payload.date,
            "type": payload.type,
            "symbol": payload.symbol,
            "quantity": str(payload.quantity),
            "price": str(payload.price),
        }
    )
    validate_incoming_transactions(connection, user, [transaction])
    trade_value = transaction["quantity"] * transaction["price"]
    cash_balance = float(user["cash_balance"])
    cash_balance = cash_balance - trade_value if transaction["type"] == "Buy" else cash_balance + trade_value

    cursor = connection.execute(
        """
        INSERT INTO transactions (user_id, date, type, symbol, quantity, price, source)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["id"],
            transaction["date"],
            transaction["type"],
            transaction["symbol"],
            transaction["quantity"],
            transaction["price"],
            "manual",
        ),
    )
    connection.execute("UPDATE users SET cash_balance = ? WHERE id = ?", (cash_balance, user["id"]))
    connection.commit()
    row = connection.execute("SELECT * FROM transactions WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return transaction_payload(row)


def generated_report_name(report_id: int, report_format: str) -> tuple[str, str]:
    normalized = report_format.lower()
    if "excel" in normalized:
        return f"financeai-report-{report_id}.xlsx", (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if "csv" in normalized:
        return f"financeai-report-{report_id}.csv", "text/csv"
    return f"financeai-report-{report_id}.pdf", "application/pdf"


def report_lines(connection: sqlite3.Connection, user: sqlite3.Row, sections: list[str]) -> list[str]:
    holdings = position_snapshot(connection, user)
    risk = risk_payload(connection, user)
    dashboard = dashboard_payload(connection, user)
    lines = [
        "FinanceAI Portfolio Report",
        f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Investor goal: {user['primary_goal']}",
        f"Risk preference: {user['risk_preference']}",
        f"Total value: {currency(total_portfolio_value(holdings, user))}",
        f"Overall risk: {risk['score']['overall']}/100 ({risk['score']['label']})",
        "",
    ]
    if not sections or "Current Holdings" in sections or "Portfolio Summary" in sections:
        lines.append("Holdings")
        for item in holdings:
            lines.append(
                f"- {item['symbol']} | {item['company']} | {currency(item['market_value'])} | "
                f"Gain/Loss {currency(item['gain_loss'])}"
            )
    if not sections or "Risk Analysis" in sections:
        lines.extend(["", "Risk Alerts"])
        for alert in risk["alerts"]:
            lines.append(f"- {alert['title']}: {alert['message']}")
    if not sections or "Performance Analysis" in sections:
        lines.extend(["", "Allocation"])
        for item in dashboard["allocation"]:
            lines.append(f"- {item['label']}: {item['value']}%")
    lines.extend(["", AI_DISCLAIMER])
    return lines


def write_pdf_report(path: Path, lines: list[str]) -> None:
    if canvas is None or letter is None:
        path.write_text("\n".join(lines), encoding="utf-8")
        return

    pdf = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    y = height - 54
    for line in lines:
        if y < 54:
            pdf.showPage()
            y = height - 54
        pdf.drawString(54, y, line[:105])
        y -= 18
    pdf.save()


def write_spreadsheet_report(path: Path, lines: list[str]) -> None:
    if Workbook is None:
        path.write_text("\n".join(lines), encoding="utf-8")
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FinanceAI Report"
    for index, line in enumerate(lines, start=1):
        worksheet.cell(row=index, column=1, value=line)
    workbook.save(path)


def create_report_file(
    connection: sqlite3.Connection,
    report_id: int,
    user: sqlite3.Row,
    report_format: str,
    sections: list[str],
) -> tuple[str, str]:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename, content_type = generated_report_name(report_id, report_format)
    path = REPORTS_DIR / filename
    lines = report_lines(connection, user, sections)

    if filename.endswith(".xlsx"):
        write_spreadsheet_report(path, lines)
    elif filename.endswith(".csv"):
        with path.open("w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            for line in lines:
                writer.writerow([line])
    else:
        write_pdf_report(path, lines)

    return str(path), content_type


def send_report_email(to_email: str, report_name: str) -> str:
    if not RESEND_API_KEY:
        return "queued_log_only"

    body = json.dumps(
        {
            "from": REPORT_FROM_EMAIL,
            "to": [to_email],
            "subject": f"Your FinanceAI report: {report_name}",
            "html": f"<p>{report_name} is ready in your FinanceAI reports page.</p>",
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        "https://api.resend.com/emails",
        data=body,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=15):
            return "sent"
    except (urllib.error.URLError, TimeoutError):
        return "queued_after_provider_error"


def seed_demo_user(connection: sqlite3.Connection) -> None:
    existing = connection.execute("SELECT id FROM users WHERE email = ?", ("profile@financeai.app",)).fetchone()
    if existing:
        return

    salt, password_hash = make_password_fields("demo1234")
    cursor = connection.execute(
        """
        INSERT INTO users (
            full_name, email, password_salt, password_hash, plan, member_since,
            risk_preference, primary_goal, base_currency, reports, cash_balance, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Portfolio User",
            "profile@financeai.app",
            salt,
            password_hash,
            "Starter",
            "March 2026",
            "Moderate",
            "Long-term growth",
            "USD",
            "Monthly summary enabled",
            6240.00,
            utc_now_iso(),
        ),
    )
    user_id = cursor.lastrowid

    for transaction in DEFAULT_TRANSACTIONS:
        connection.execute(
            """
            INSERT INTO transactions (user_id, date, type, symbol, quantity, price, source)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                transaction["date"],
                transaction["type"],
                transaction["symbol"],
                transaction["quantity"],
                transaction["price"],
                "starter_portfolio.csv",
            ),
        )

    connection.execute(
        "INSERT INTO watchlist (user_id, symbol, created_at) VALUES (?, ?, ?)",
        (user_id, "AMD", utc_now_iso()),
    )
    connection.execute(
        """
        INSERT INTO uploads (user_id, name, date_label, status, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (user_id, "starter_portfolio.csv", "Seeded for demo", "Processed", utc_now_iso()),
    )
    connection.execute(
        """
        INSERT INTO reports (user_id, title, meta, report_format, date_range, sections, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            "Portfolio Report - February 2026",
            "2026-03-01 | 1.4 MB",
            "PDF Document",
            "Last Month",
            "Portfolio Summary,Performance Analysis,Current Holdings",
            utc_now_iso(),
        ),
    )
    connection.execute(
        """
        INSERT INTO reports (user_id, title, meta, report_format, date_range, sections, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            "Quarterly Snapshot - Q1 2026",
            "2026-03-21 | 2.1 MB",
            "PDF Document",
            "Last Quarter",
            "Portfolio Summary,Risk Analysis,Current Holdings",
            utc_now_iso(),
        ),
    )
    connection.commit()


def init_db() -> None:
    with get_db() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                plan TEXT NOT NULL DEFAULT 'Starter',
                member_since TEXT NOT NULL,
                risk_preference TEXT NOT NULL DEFAULT 'Moderate',
                primary_goal TEXT NOT NULL DEFAULT 'Long-term growth',
                base_currency TEXT NOT NULL DEFAULT 'USD',
                reports TEXT NOT NULL DEFAULT 'Monthly summary enabled',
                cash_balance REAL NOT NULL DEFAULT 5000,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                type TEXT NOT NULL,
                symbol TEXT NOT NULL,
                quantity REAL NOT NULL,
                price REAL NOT NULL,
                source TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS watchlist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                symbol TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, symbol),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                date_label TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                meta TEXT NOT NULL,
                report_format TEXT,
                date_range TEXT,
                sections TEXT,
                file_path TEXT,
                content_type TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS report_emails (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                email TEXT NOT NULL,
                report_name TEXT NOT NULL,
                queued_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            """
        )
        for statement in [
            "ALTER TABLE reports ADD COLUMN file_path TEXT",
            "ALTER TABLE reports ADD COLUMN content_type TEXT",
        ]:
            try:
                connection.execute(statement)
            except sqlite3.OperationalError:
                pass
        if SEED_DEMO_DATA:
            seed_demo_user(connection)


@app.on_event("startup")
def startup_event() -> None:
    init_db()


@app.get("/")
def root() -> dict[str, str]:
    return {"message": "FinanceAI API is running.", "docs": "/docs"}


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "timestamp": utc_now_iso(),
        "database": str(DB_PATH),
        "ai_provider": "openai" if OPENAI_API_KEY else "local_rules",
        "allowed_origins": ALLOWED_ORIGINS,
    }


@app.post("/api/auth/signup")
def signup(payload: SignupRequest) -> dict[str, Any]:
    with get_db() as connection:
        existing = connection.execute("SELECT id FROM users WHERE email = ?", (payload.email,)).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="An account with this email already exists.")

        salt, password_hash = make_password_fields(payload.password)
        cursor = connection.execute(
            """
            INSERT INTO users (
                full_name, email, password_salt, password_hash, plan, member_since,
                risk_preference, primary_goal, base_currency, reports, cash_balance, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.full_name,
                payload.email,
                salt,
                password_hash,
                "Starter",
                datetime.now(UTC).strftime("%B %Y"),
                "Moderate",
                payload.goal,
                "USD",
                "Monthly summary enabled",
                10000.00,
                utc_now_iso(),
            ),
        )
        user_id = cursor.lastrowid
        connection.execute(
            """
            INSERT INTO uploads (user_id, name, date_label, status, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (user_id, "welcome_template.csv", "Ready for your first import", "Pending", utc_now_iso()),
        )
        
        if SEED_DEMO_DATA:
            for txn in DEFAULT_TRANSACTIONS:
                connection.execute(
                    """
                    INSERT INTO transactions (user_id, date, type, symbol, quantity, price, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (user_id, txn["date"], txn["type"], txn["symbol"], txn["quantity"], txn["price"], utc_now_iso()),
                )

        token = create_session(connection, user_id)
        user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    return {"message": "Signup successful", "user": public_user_profile(user), "token": token}


@app.post("/api/auth/login")
def login(payload: LoginRequest) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE email = ?", (payload.email,)).fetchone()
        if not user or not verify_password(payload.password, user["password_salt"], user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid email or password.")
        token = create_session(connection, user["id"])

    return {"message": "Login successful", "user": public_user_profile(user), "token": token}


@app.get("/api/auth/me")
def auth_me(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    return {"user": public_user_profile(session["user"])}


@app.post("/api/auth/logout")
def logout(session: dict[str, Any] = Depends(current_user)) -> dict[str, str]:
    with get_db() as connection:
        connection.execute("DELETE FROM sessions WHERE token = ?", (session["token"],))
        connection.commit()
    return {"message": "Logged out successfully"}


@app.get("/api/profile/me")
def get_profile(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    return public_user_profile(session["user"])


@app.put("/api/profile/me")
def update_profile(payload: ProfileUpdateRequest, session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        duplicate = connection.execute(
            "SELECT id FROM users WHERE email = ? AND id != ?",
            (payload.email, session["user"]["id"]),
        ).fetchone()
        if duplicate:
            raise HTTPException(status_code=409, detail="That email is already in use.")

        connection.execute(
            """
            UPDATE users
            SET full_name = ?, email = ?, risk_preference = ?, primary_goal = ?, base_currency = ?, reports = ?
            WHERE id = ?
            """,
            (
                payload.full_name,
                payload.email,
                payload.risk_preference,
                payload.primary_goal,
                payload.base_currency,
                payload.reports,
                session["user"]["id"],
            ),
        )
        connection.commit()
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()

    return {"message": "Profile updated", "profile": public_user_profile(user)}


@app.get("/api/dashboard")
def get_dashboard(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return dashboard_payload(connection, user)


@app.get("/api/uploads")
def get_upload_page_data(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        return uploads_payload(connection, session["user"]["id"])


@app.post("/api/uploads/transactions")
async def upload_transactions(
    file: UploadFile = File(...),
    session: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    filename = file.filename or "upload"
    contents = await file.read()

    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()

        if filename.lower().endswith(".csv"):
            decoded = contents.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
            if not rows:
                raise HTTPException(status_code=400, detail="CSV file is empty.")
            try:
                parsed_rows = [coerce_transaction(row) for row in rows]
                return insert_transactions(connection, user, parsed_rows, filename)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc)) from exc

        if filename.lower().endswith((".xlsx", ".xls")):
            try:
                parsed_rows = parse_excel_transactions(contents)
                return insert_transactions(connection, user, parsed_rows, filename)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc)) from exc

    raise HTTPException(status_code=400, detail="Only CSV, XLSX, and XLS files are supported.")


@app.get("/api/transactions")
def get_transactions(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        rows = connection.execute(
            """
            SELECT *
            FROM transactions
            WHERE user_id = ?
            ORDER BY date DESC, id DESC
            """,
            (session["user"]["id"],),
        ).fetchall()
        return {"transactions": [transaction_payload(row) for row in rows]}


@app.post("/api/transactions")
def create_transaction(
    payload: TransactionRequest,
    session: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        try:
            transaction = insert_single_transaction(connection, user, payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"message": "Transaction added and portfolio refreshed.", "transaction": transaction}


@app.delete("/api/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, session: dict[str, Any] = Depends(current_user)) -> dict[str, str]:
    with get_db() as connection:
        row = connection.execute(
            "SELECT * FROM transactions WHERE id = ? AND user_id = ?",
            (transaction_id, session["user"]["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Transaction not found.")

        trade_value = float(row["quantity"]) * float(row["price"])
        cash_delta = trade_value if row["type"] == "Buy" else -trade_value
        connection.execute(
            "UPDATE users SET cash_balance = cash_balance + ? WHERE id = ?",
            (cash_delta, session["user"]["id"]),
        )
        connection.execute("DELETE FROM transactions WHERE id = ?", (transaction_id,))
        connection.commit()
    return {"message": "Transaction deleted and cash balance adjusted."}


@app.get("/api/risk")
def get_risk_analysis(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return risk_payload(connection, user)


@app.get("/api/recommendations")
def get_recommendations(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return recommendations_payload(connection, user)


@app.post("/api/recommendations/{ticker}/watchlist")
def add_to_watchlist(ticker: str, session: dict[str, Any] = Depends(current_user)) -> dict[str, str]:
    symbol = ticker.upper()
    with get_db() as connection:
        existing = connection.execute(
            "SELECT id FROM watchlist WHERE user_id = ? AND symbol = ?",
            (session["user"]["id"], symbol),
        ).fetchone()
        if existing:
            return {"message": f"{symbol} is already on your watchlist"}

        connection.execute(
            "INSERT INTO watchlist (user_id, symbol, created_at) VALUES (?, ?, ?)",
            (session["user"]["id"], symbol, utc_now_iso()),
        )
        connection.commit()
    return {"message": f"{symbol} added to watchlist"}


@app.post("/api/ai/portfolio-summary")
def ai_portfolio_summary(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return call_openai_insight("portfolio_summary", ai_context(connection, user))


@app.post("/api/ai/risk-explanation")
def ai_risk_explanation(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return call_openai_insight("risk_explanation", ai_context(connection, user))


@app.post("/api/ai/recommendations")
def ai_recommendation_explanation(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return call_openai_insight("recommendation_explanation", ai_context(connection, user))


@app.post("/api/ai/report-insights")
def ai_report_insights(
    payload: AIReportInsightRequest,
    session: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        context = ai_context(connection, user)
        context["report_request"] = {
            "format": payload.format,
            "date_range": payload.date_range,
            "sections": payload.sections,
        }
        return call_openai_insight("report_insights", context)


@app.post("/api/ai/chat")
def ai_chat(payload: AIChatRequest, session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        return call_openai_insight("chat", ai_context(connection, user), payload.question)


@app.get("/api/reports")
def get_reports(session: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with get_db() as connection:
        return reports_payload(connection, session["user"]["id"])


@app.post("/api/reports/generate")
def generate_report(
    payload: ReportGenerateRequest,
    session: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    title = f"{payload.format} - {payload.date_range}"
    meta = f"{datetime.now().date().isoformat()} | {max(1, len(payload.sections)) * 0.2:.1f} MB"
    with get_db() as connection:
        cursor = connection.execute(
            """
            INSERT INTO reports (user_id, title, meta, report_format, date_range, sections, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session["user"]["id"],
                title,
                meta,
                payload.format,
                payload.date_range,
                ",".join(payload.sections),
                utc_now_iso(),
            ),
        )
        report_id = cursor.lastrowid
        user = connection.execute("SELECT * FROM users WHERE id = ?", (session["user"]["id"],)).fetchone()
        file_path, content_type = create_report_file(
            connection,
            report_id,
            user,
            payload.format,
            payload.sections,
        )
        connection.execute(
            "UPDATE reports SET file_path = ?, content_type = ? WHERE id = ?",
            (file_path, content_type, report_id),
        )
        connection.commit()
    return {
        "message": f"Report generated with {len(payload.sections)} sections.",
        "report": {"id": report_id, "title": title, "meta": meta, "content_type": content_type},
        "format": payload.format,
        "sections": payload.sections,
    }


@app.get("/api/reports/{report_id}/download")
def download_report(report_id: int, session: dict[str, Any] = Depends(current_user)) -> FileResponse:
    with get_db() as connection:
        report = connection.execute(
            """
            SELECT title, file_path, content_type
            FROM reports
            WHERE id = ? AND user_id = ?
            """,
            (report_id, session["user"]["id"]),
        ).fetchone()
    if not report or not report["file_path"]:
        raise HTTPException(status_code=404, detail="Report file not found.")

    path = Path(report["file_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="Report file is no longer available.")

    return FileResponse(
        path=path,
        media_type=report["content_type"] or "application/octet-stream",
        filename=path.name,
    )


@app.post("/api/reports/email")
def email_report(payload: EmailReportRequest, session: dict[str, Any] = Depends(current_user)) -> dict[str, str]:
    delivery_status = send_report_email(payload.email, payload.report_name)
    with get_db() as connection:
        connection.execute(
            """
            INSERT INTO report_emails (user_id, email, report_name, queued_at)
            VALUES (?, ?, ?, ?)
            """,
            (session["user"]["id"], payload.email, payload.report_name, utc_now_iso()),
        )
        connection.commit()
    if delivery_status == "sent":
        return {"message": f"{payload.report_name} sent to {payload.email}"}
    return {"message": f"{payload.report_name} queued for delivery to {payload.email} ({delivery_status})"}
